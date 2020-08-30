import pandas as pd
import re
from word2number import w2n
import xlrd 
import numbers
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, colors
import numpy as np
import math

# missing_values_list = ["n/a", "na", "--","NaN", "nan","-"]
# data = pd.read_excel("C:\\Users\\DELL\\Documents\\project\\one.xlsx",na_values = missing_values_list)
# # Parameters printing
# print("The parameters of your data set are: ")
# for col in data.columns: 
#     print(col) 
# p=list(data.columns) 
# converting dataframe to str and then lower case

def conversion(data):
    data1=pd.concat([data[col].astype(str).str.lower() for col in data.columns], axis=1)
    return data1
# removing extra spaces in the start and end 
# allowing one space in btw 
# removing all spaces in btw for int cols

def spaces(data1):
    datat = data1.apply(lambda x: x.str.strip() if x.dtype == "object" else x)
    datat = data1.apply(lambda x: x.str.replace(" ","") if x.dtype == "object" else x)
    #datat['name']=datat['name'].apply(lambda x: x.str.replace('\s+', ' ', regex=True) if x.dtype == "object" else x)
    return datat
# converting words to num
# converting the int cols back to int from str

def wordtonum(data):
    z=[]
    for i in range (0,len(data.columns)):
        c = 0
        for j in range(0,len(data)):
            if type(data.iat[j,i])==int:
                c=c+1
            else:
                continue
        if c>=(len(data)/2):
            z.append(i)
            # for k in range (0,len(data)):
            #     if type(data.iat[k,i]) == str :
            #         try:
            #             data.iat[k,i]=w2n.word_to_num(data.iat[k,i])
            #         except ValueError :
            #             pass
            print(z)
    return data,z

# converting words to num
# converting the int cols back to int from str
def wordtonum1(datat,z):
    for i in z:
        for k in range (0,len(datat)):
            if type(datat.iat[k,i]) == str :
                try:
                    datat.iat[k,i]=w2n.word_to_num(datat.iat[k,i])
                except ValueError :
                    pass
    return datat

def duplicates_removal(d):
    d.drop_duplicates(keep = "first", inplace = True) 
    return d
# no applying for name col
# spell check

def spellcheck(datat):
    f=open("C:/Users/Rakesh/Desktop/CS Project/App/Error_Report/Error.txt","w+") #opening a new text file for reporting spelling errors
    string=""
    f.write("\n")
    set1=[]
    workbook = load_workbook(filename="C:/Users/Rakesh/Desktop/CS Project/App/Upload_Folder/two.xlsx")
    sheet = workbook.active
    for i in range(0,len(datat)):
        for j in range(0,len(datat.columns)):
            if type(datat.iat[i,j]) == str  :
                if len(datat.iat[i,j].split()) == 1:
                    set1.append(datat.iat[i,j])
               # elif len(sheet.cell_value(i,j).split()) > 1:
    
    words = open("C:/Users/Rakesh/Desktop/CS Project/App/static/Additional/spell_words.txt").readlines()
    words = [word.strip() for word in words]
    d=[]
    for i in range(0,len(set1)):
            if set1[i] not in words:
                d.append(set1[i])
    #print(d)
    #print("SPELLING MISTAKES IN: ")
    for i in range(0,len(d)):
        for j in range(0,len(datat)):
            for k in range(0,len(datat.columns)):
                if d[i]==datat.iat[j,k]:
                    #print(j+2,"row and",k+1,"column")
                    string="Spelling Error at row " +str(j+2) +" column "+str(k+1)+" \n"
                    f.write(string) # Writing Findings into text file
                    sheet.cell(row=j+2, column=k+1).fill = PatternFill(fgColor='FFEE08', fill_type = "solid")
    workbook.save(filename="final.xlsx")
    f.write('------------------------------------------------------------------------------------------------------------\n')
    f.close()
def missing_summary(datat):
    f=open("C:/Users/Rakesh/Desktop/CS Project/App/Error_Report/Error_summary.txt","a") #opening the existing text file for reporting missing values
    string=""
    f.write('\n')
    missing_values_count=datat.isnull().sum()
    summary=missing_values_count[0:100]
    total_cells=np.product(datat.shape)
    
    # percent of data that is missing
    total_missing = missing_values_count.sum()
    percent_miss=(total_missing/total_cells) * 100
    percent_miss=math.ceil(percent_miss)
    f.write("Percentage of Missing Data: ")
    f.write(str(percent_miss))
    f.write('%')
    f.write('\n')
    f.write(str(summary))
    f.close()
def missing_values(datat):
    f=open("C:/Users/Rakesh/Desktop/CS Project/App/Error_Report/Error.txt","a") #opening the existing text file for reporting missing values
    string=""
    workbook = load_workbook(filename="C:/Users/Rakesh/Desktop/CS Project/App/final.xlsx")
    sheet = workbook.active
    missing_values_list = ["n/a", "na", "--","NaN", "nan","-"]
#print("\nMISSING VALUES IN:")
    for i in range(0,len(missing_values_list)):
        for j in range(0,len(datat)):
            for k in range(0,len(datat.columns)):
                if missing_values_list[i]==datat.iat[j,k]:
                    #print(j+2,"row and",k+1,"column")
                    string="Missing value in Row "+str(j+2)+" column "+str(k+1)+" \n"
                    f.write(string) # Writing Findings into text file
                    sheet.cell(row=j+2, column=k+1).fill = PatternFill(fgColor='FF0000', fill_type = "solid")
                    try:
                        median = datat.iloc[:,k].median()
                        sheet.cell(row=j+2,column=k+1).value= median
                        datat.iat[j,k]=median
                    except(TypeError):
                        continue
    workbook.save(filename="final.xlsx")
    f.close()
def fix_text():
    lines_seen=set()
    outfile=open("C:/Users/Rakesh/Desktop/CS Project/App/Error_Report/Error_final.txt","w")
    for line in open("C:/Users/Rakesh/Desktop/CS Project/App/Error_Report/Error.txt","r"):
        if line not in lines_seen:
            outfile.write(line)
            lines_seen.add(line)
    outfile.close()
    fd=open("C:/Users/Rakesh/Desktop/CS Project/App/Error_Report/Error_summary.txt","r")
    d=fd.read()
    fd.close()
    m=d.split("\n")
    s="\n".join(m[:-1])
    fd=open("C:/Users/Rakesh/Desktop/CS Project/App/Error_Report/Error_summary_final.txt","w+")
    for i in range(len(s)):
        fd.write(s[i])
    fd.close() 
def run(data):
    missing_summary(data)
    a,z=wordtonum(data)
    b=conversion(a)
    c=spaces(b)
    d=wordtonum1(c,z)
    e=duplicates_removal(d)
    e.to_excel('C:/Users/Rakesh/Desktop/CS Project/App/Upload_Folder/two.xlsx',index=False)
    spellcheck(e)
    missing_values(e)
    fix_text()
